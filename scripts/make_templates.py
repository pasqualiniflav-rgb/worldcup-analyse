"""Génère les classeurs Excel de SAISIE (interface du staff).

Produit deux fichiers dans data/ :
  - data/templates/modele_saisie.xlsx  -> gabarit vierge à distribuer
  - data/input/saisie_selection.xlsx   -> le même, pré-rempli de données
                                          fictives pour la démo

Le staff ne travaille QUE dans ce type de classeur. Les listes
déroulantes viennent de config/client.yaml pour cadrer la saisie.
"""
from __future__ import annotations

import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from config import load_config, TEMPLATES_DIR, INPUT_DIR

CFG = load_config()
PRIM = CFG["branding"]["couleur_primaire"].lstrip("#")
SEC = CFG["branding"]["couleur_secondaire"].lstrip("#")

HEADER_FILL = PatternFill("solid", fgColor=PRIM)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=PRIM, bold=True, size=14)
HINT_FONT = Font(color="808080", italic=True, size=9)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Définition des feuilles : (nom, titre, [(colonne, indice, largeur)])
SHEETS = {
    "Joueurs": {
        "titre": "MODULE JOUEURS — Fiche d'identité",
        "cols": [
            ("player_id", "ex. JOU001 (unique, ne pas réutiliser)", 12),
            ("nom", "NOM de famille", 16),
            ("prenom", "Prénom", 16),
            ("date_naissance", "AAAA-MM-JJ", 14),
            ("poste", "liste", 10),
            ("pied", "Droit / Gauche / Ambidextre", 14),
            ("taille_cm", "en cm", 10),
            ("selection", "liste", 14),
            ("statut", "liste", 14),
            ("club", "Club actuel", 20),
            ("pays_club", "Pays du club", 14),
            ("notes_generales", "Profil, points forts/faibles...", 40),
        ],
    },
    "Observations": {
        "titre": "MODULE JOUEURS — Observations vidéo / match",
        "cols": [
            ("player_id", "doit exister dans l'onglet Joueurs", 12),
            ("date", "AAAA-MM-JJ", 14),
            ("contexte", "liste", 12),
            ("competition", "ex. Ligue 1, Éliminatoires", 20),
            ("adversaire", "Équipe adverse ce jour-là", 18),
            ("minutes", "min. observées / jouées", 12),
            ("note_globale", "note sur l'échelle configurée", 14),
            ("video_url", "lien YouTube / fichier", 34),
            ("video_ts", "horodatage en SECONDES", 14),
            ("code_action", "ex. But, Passe D, Duel gagné", 18),
            ("commentaire", "observation détaillée", 40),
            ("observateur", "qui a codé", 14),
        ],
    },
    "Medical": {
        "titre": "MODULE MÉDICAL / DISPONIBILITÉ",
        "cols": [
            ("player_id", "doit exister dans Joueurs", 12),
            ("date", "AAAA-MM-JJ", 14),
            ("statut", "liste", 14),
            ("type_blessure", "ex. Ischio, Cheville", 18),
            ("retour_prevu", "AAAA-MM-JJ estimé", 16),
            ("commentaire", "détail", 40),
        ],
    },
    "TempsJeu": {
        "titre": "TEMPS DE JEU RÉEL EN CLUB",
        "cols": [
            ("player_id", "doit exister dans Joueurs", 12),
            ("date", "AAAA-MM-JJ", 14),
            ("club", "Club", 18),
            ("competition", "Compétition", 18),
            ("minutes", "minutes jouées", 12),
            ("titulaire", "1 = titulaire, 0 = entré/banc", 14),
            ("buts", "buts marqués", 10),
            ("passes_d", "passes décisives", 12),
        ],
    },
    "Adversaires": {
        "titre": "MODULE ADVERSAIRE — Préparation match international",
        "cols": [
            ("adversaire_id", "ex. ADV_ESP", 12),
            ("pays", "Pays adverse", 16),
            ("systeme", "ex. 4-3-3", 12),
            ("forces", "points forts", 34),
            ("faiblesses", "points faibles", 34),
            ("joueurs_cles", "joueurs à surveiller", 30),
            ("video_url", "lien vidéo", 30),
            ("commentaire", "plan de jeu", 34),
        ],
    },
}

# Listes déroulantes par colonne (depuis la config)
DROPDOWNS = {
    "poste": CFG["categories"]["postes"],
    "selection": CFG["categories"]["selections"],
    "statut": CFG["categories"]["statuts_convocation"],
    "contexte": ["Sélection", "Club"],
    ("Medical", "statut"): ["Apte", "Blessé", "Reprise", "Incertain"],
}


def _style_sheet(ws, sheet_name, spec):
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(spec["cols"]))
    tcell = ws.cell(row=1, column=1, value=spec["titre"])
    tcell.font = TITLE_FONT
    # Ligne d'en-têtes (row 3) + ligne d'aide (row 4)
    for j, (col, hint, width) in enumerate(spec["cols"], start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = width
        h = ws.cell(row=3, column=j, value=col)
        h.fill = HEADER_FILL
        h.font = HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = BORDER
        hint_cell = ws.cell(row=4, column=j, value=hint)
        hint_cell.font = HINT_FONT
        hint_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A5"  # données à partir de la ligne 5

    # Listes déroulantes
    for j, (col, hint, width) in enumerate(spec["cols"], start=1):
        options = DROPDOWNS.get((sheet_name, col)) or DROPDOWNS.get(col)
        if options:
            letter = get_column_letter(j)
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(options) + '"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(f"{letter}5:{letter}500")


def _write_rows(ws, spec, rows):
    keys = [c[0] for c in spec["cols"]]
    for i, row in enumerate(rows, start=5):
        for j, key in enumerate(keys, start=1):
            c = ws.cell(row=i, column=j, value=row.get(key, ""))
            c.border = BORDER


def build_workbook(data=None):
    wb = Workbook()
    wb.remove(wb.active)
    for name, spec in SHEETS.items():
        ws = wb.create_sheet(title=name)
        _style_sheet(ws, name, spec)
        if data and name in data:
            _write_rows(ws, spec, data[name])
    return wb


# --------------------------------------------------------------------------
#  Données fictives d'une sélection nationale (démo)
# --------------------------------------------------------------------------
def _fake_data():
    random.seed(42)  # reproductible

    joueurs = [
        dict(player_id="JOU001", nom="Konaté", prenom="Ibrahim", date_naissance="1998-04-12",
             poste="DC", pied="Droit", taille_cm=189, selection="A", statut="Cadre",
             club="Olympique Lyonnais", pays_club="France",
             notes_generales="Défenseur axial dominant dans les airs, relance propre. Leader."),
        dict(player_id="JOU002", nom="Mendes", prenom="Rafael", date_naissance="2001-09-03",
             poste="AILG", pied="Droit", taille_cm=176, selection="A", statut="Cadre",
             club="FC Porto", pays_club="Portugal",
             notes_generales="Ailier percutant, un contre un tranchant. À fiabiliser défensivement."),
        dict(player_id="JOU003", nom="Traoré", prenom="Sekou", date_naissance="2003-01-22",
             poste="MC", pied="Gauche", taille_cm=181, selection="Espoirs U21", statut="À suivre",
             club="RC Lens", pays_club="France",
             notes_generales="Milieu box-to-box, gros volume. Prise de décision à affiner."),
        dict(player_id="JOU004", nom="Bianchi", prenom="Marco", date_naissance="1996-11-30",
             poste="GB", pied="Droit", taille_cm=191, selection="A", statut="Rotation",
             club="Torino FC", pays_club="Italie",
             notes_generales="Gardien serein sur sa ligne, jeu au pied correct."),
        dict(player_id="JOU005", nom="Nakamura", prenom="Yuto", date_naissance="2002-06-18",
             poste="MOC", pied="Droit", taille_cm=173, selection="Espoirs U21", statut="À suivre",
             club="RSC Anderlecht", pays_club="Belgique",
             notes_generales="Meneur créatif, vision et dernière passe. Physique à renforcer."),
        dict(player_id="JOU006", nom="Silva", prenom="Diego", date_naissance="1999-02-14",
             poste="BU", pied="Droit", taille_cm=185, selection="A", statut="Cadre",
             club="Villarreal CF", pays_club="Espagne",
             notes_generales="Avant-centre de surface, bon jeu dos au but. Réaliste."),
        dict(player_id="JOU007", nom="Diallo", prenom="Amadou", date_naissance="2004-03-09",
             poste="DD", pied="Droit", taille_cm=178, selection="U19", statut="Vivier",
             club="Centre de formation FN", pays_club="France",
             notes_generales="Latéral moderne, projections. Grosse marge de progression."),
        dict(player_id="JOU008", nom="Petrov", prenom="Nikola", date_naissance="1997-07-25",
             poste="MDF", pied="Droit", taille_cm=183, selection="A", statut="Rotation",
             club="PSV Eindhoven", pays_club="Pays-Bas",
             notes_generales="Sentinelle, lecture des trajectoires, récupération haute."),
        dict(player_id="JOU009", nom="Rossi", prenom="Luca", date_naissance="2005-10-01",
             poste="AVC", pied="Gauche", taille_cm=180, selection="U17", statut="Vivier",
             club="Centre de formation FN", pays_club="France",
             notes_generales="Grand espoir offensif, finition des deux pieds. À protéger."),
        dict(player_id="JOU010", nom="Sow", prenom="Cheikh", date_naissance="2000-12-11",
             poste="DG", pied="Gauche", taille_cm=180, selection="A", statut="À suivre",
             club="Stade Rennais", pays_club="France",
             notes_generales="Latéral gauche endurant. Centres à améliorer."),
    ]

    codes = ["But", "Passe D", "Duel gagné", "Récupération", "Occasion manquée",
             "Erreur technique", "Séquence pressing", "Ouverture"]
    competitions = ["Ligue 1", "Liga", "Serie A", "Eredivisie", "Éliminatoires CM",
                    "Ligue des Nations", "Amical"]
    advs = ["Espagne", "Croatie", "Sénégal", "OM", "Real Madrid", "Ajax", "Inter"]
    observateurs = ["P. Garnier", "M. Lefèvre", "A. Costa"]

    today = date(2026, 7, 20)

    observations = []
    for jou in joueurs:
        for k in range(random.randint(2, 4)):
            d = today - timedelta(days=random.randint(3, 120))
            ctx = random.choice(["Sélection", "Club", "Club"])
            observations.append(dict(
                player_id=jou["player_id"],
                date=d.isoformat(),
                contexte=ctx,
                competition=random.choice(competitions),
                adversaire=random.choice(advs),
                minutes=random.choice([45, 60, 70, 90, 90, 90]),
                note_globale=round(random.uniform(4.5, 8.5), 1),
                video_url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
                video_ts=random.randint(120, 5400),
                code_action=random.choice(codes),
                commentaire=random.choice([
                    "Bon repli défensif, ferme l'intervalle.",
                    "Prise d'info avant réception, casse une ligne.",
                    "Perd son duel, en retard sur l'appui.",
                    "Frappe cadrée sur corner, présence au deuxième poteau.",
                    "Déchet technique dans le dernier tiers.",
                    "Leadership vocal, replace ses partenaires.",
                ]),
                observateur=random.choice(observateurs),
            ))

    medical = [
        dict(player_id="JOU002", date=(today - timedelta(days=8)).isoformat(),
             statut="Blessé", type_blessure="Ischio (grade 1)",
             retour_prevu=(today + timedelta(days=10)).isoformat(),
             commentaire="Ressenti à l'entraînement club, à réévaluer avant convocation."),
        dict(player_id="JOU006", date=(today - timedelta(days=2)).isoformat(),
             statut="Apte", type_blessure="", retour_prevu="",
             commentaire="RAS, disponible."),
        dict(player_id="JOU008", date=(today - timedelta(days=5)).isoformat(),
             statut="Reprise", type_blessure="Cheville",
             retour_prevu=(today + timedelta(days=3)).isoformat(),
             commentaire="Reprise partielle collective, sensations bonnes."),
        dict(player_id="JOU005", date=(today - timedelta(days=1)).isoformat(),
             statut="Incertain", type_blessure="Adducteurs",
             retour_prevu=(today + timedelta(days=6)).isoformat(),
             commentaire="Gêne persistante, décision J-3."),
    ]

    temps_jeu = []
    for jou in joueurs:
        for k in range(random.randint(4, 8)):
            d = today - timedelta(days=random.randint(2, 90))
            titu = random.random() > 0.3
            mins = random.choice([90, 90, 90, 75, 60, 20, 0]) if titu else random.choice([0, 15, 25, 30])
            temps_jeu.append(dict(
                player_id=jou["player_id"],
                date=d.isoformat(),
                club=jou["club"],
                competition=random.choice(competitions),
                minutes=mins,
                titulaire=1 if titu else 0,
                buts=random.choice([0, 0, 0, 1]) if jou["poste"] in ("BU", "AVC", "AILG", "MOC") else 0,
                passes_d=random.choice([0, 0, 1]),
            ))

    adversaires = [
        dict(adversaire_id="ADV_ESP", pays="Espagne", systeme="4-3-3",
             forces="Possession, jeu de position, densité au milieu.",
             faiblesses="Vulnérable dans le dos des latéraux, transitions rapides.",
             joueurs_cles="N°8 relayeur, ailier gauche 1v1",
             video_url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
             commentaire="Bloquer les demi-espaces, attaquer la profondeur en transition."),
        dict(adversaire_id="ADV_CRO", pays="Croatie", systeme="4-3-3",
             forces="Milieu technique, maîtrise du tempo.",
             faiblesses="Vitesse défensive limitée, seconde période.",
             joueurs_cles="Meneur reculé, avant-centre pivot",
             video_url="https://www.youtube.com/watch?v=dQw4w9WgXcQ",
             commentaire="Presser le milieu, jouer les intervalles côté droit."),
    ]

    return dict(Joueurs=joueurs, Observations=observations, Medical=medical,
                TempsJeu=temps_jeu, Adversaires=adversaires)


def main():
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1) Gabarit vierge
    build_workbook().save(TEMPLATES_DIR / "modele_saisie.xlsx")

    # 2) Classeur pré-rempli (démo)
    build_workbook(_fake_data()).save(INPUT_DIR / "saisie_selection.xlsx")

    print("Gabarit vierge :", TEMPLATES_DIR / "modele_saisie.xlsx")
    print("Démo remplie   :", INPUT_DIR / "saisie_selection.xlsx")


if __name__ == "__main__":
    main()
